import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# === Nastavení cest ===
base_path = os.path.dirname(os.path.abspath(__file__))
original_file = [f for f in os.listdir(base_path) if f.endswith(".xlsx") and not f.startswith("~$")][0]
original_path = os.path.join(base_path, original_file)
import_dir = os.path.join(base_path, "ImportNewData")
backup_dir = os.path.join(base_path, "DataBackup")
os.makedirs(backup_dir, exist_ok=True)

# === Funkce pro načtení tabulky DataTable ===
def load_datatable_from_file(file_path, table_name="DataTable"):
    wb = load_workbook(file_path, data_only=True)
    for ws in wb.worksheets:
        if table_name in ws.tables:
            ref = ws.tables[table_name].ref
            data = ws[ref]
            headers = [cell.value for cell in next(iter(data))]
            rows = [[cell.value for cell in row] for row in list(data)[1:]]
            df = pd.DataFrame(rows, columns=headers)
            return df
    raise ValueError(f"Tabulka '{table_name}' nebyla nalezena v souboru: {file_path}")

# === Záloha původního souboru ===
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
os.replace(original_path, backup_path)

# === Načtení původní tabulky a zjištění max Serial_Number ===
df_original = load_datatable_from_file(backup_path)
df_original.columns = [str(c).strip().replace(" ", "_") for c in df_original.columns]

if "Serial_Number" not in df_original.columns:
    raise ValueError("Chybí sloupec 'Serial_Number' v původní tabulce.")

df_original["Serial_Number"] = pd.to_numeric(df_original["Serial_Number"], errors="coerce")
max_serial = df_original["Serial_Number"].max()

# === Zpracování nových souborů ===
new_rows = []
for filename in os.listdir(import_dir):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(import_dir, filename)
        try:
            df_new = load_datatable_from_file(file_path)
            df_new.columns = [str(c).strip().replace(" ", "_") for c in df_new.columns]
            df_new["Serial_Number"] = pd.to_numeric(df_new["Serial_Number"], errors="coerce")
            df_new = df_new[df_new["Serial_Number"] > max_serial]
            if not df_new.empty:
                print(f"🆕 Přidávám nové řádky ze souboru {filename}:\n{df_new[['Serial_Number']].head()}")
                new_rows.append(df_new)
        except Exception as e:
            print(f"⚠️ Chyba při načítání {filename}: {e}")

# === Spojení a zápis nové verze ===
df_final = pd.concat([df_original] + new_rows, ignore_index=True) if new_rows else df_original.copy()

# === Otevření původního souboru a smazání listu LastVersion ===
wb = load_workbook(backup_path)
if "LastVersion" in wb.sheetnames:
    del wb["LastVersion"]
ws = wb.create_sheet("LastVersion")

# === Zapsání nových dat ===
for row in dataframe_to_rows(df_final, index=False, header=True):
    ws.append(row)

# === Vytvoření nové tabulky DataTable ===
max_row = ws.max_row
max_col = ws.max_column
end_col = get_column_letter(max_col)
table_range = f"A1:{end_col}{max_row}"

table = Table(displayName="DataTable", ref=table_range)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# === Uložení zpět do původního názvu ===
wb.save(original_path)
print("✅ Hotovo. Přidány nové řádky. Výstup je v listu 'LastVersion' jako tabulka 'DataTable'.")
